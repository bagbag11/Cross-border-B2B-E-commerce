#!/usr/bin/env python3
"""
跨境物流报价表解析器 v4
支持20种货代格式，输出对齐飞书多维表格Schema
"""
import argparse, json, os, re, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error":"缺少openpyxl依赖"})); sys.exit(1)

try:
    import xlrd
except ImportError:
    xlrd = None

# ============================================================
# 工具函数
# ============================================================

def sf(v):
    """safe_float"""
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s = re.sub(r'[^\d.\-]','',str(v).strip().replace(",","").replace("，",""))
    try: return float(s) if s else None
    except: return None

def ss(v): return "" if v is None else str(v).strip()

def cell(ws,r,c):
    try: return ws.cell(row=r,column=c).value
    except: return None

def split_countries(text):
    text = ss(text)
    if not text: return []
    for sep in ["、","，",","," / ","/","\n"]:
        if sep in text:
            parts = [p.strip() for p in text.split(sep) if p.strip()]
            if len(parts)>1: return parts
    return [text] if text else []

def detect_supplier(filename, sheet_names):
    fn = filename.lower(); sn_all = " ".join(sheet_names).lower()
    if "蓝牛" in fn: return "蓝牛国际"
    if "亿俐缇" in fn: return "亿俐缇"
    if "福鑫" in fn: return "福鑫国际"
    if "韩润" in fn: return "韩润物流"
    if "祥源" in fn: return "祥源国际"
    if "畅骏" in fn: return "畅骏国际"
    if "启丰" in fn: return "启丰物流"
    if "海源" in fn: return "海源日本"
    if "卉马" in fn: return "卉马国际"
    if "昌海" in fn: return "昌海运通"
    if "翔平" in fn: return "翔平"
    if "艾姆勒" in fn or "iml" in fn.lower(): return "艾姆勒"
    if "邮政大包" in fn and "惠运捷" in fn: return "惠运捷邮政大包"
    if "满天星" in fn: return "满天星综合" if "综合" in fn else "满天星主营"
    if "惠运捷" in fn: return "惠运捷"
    if "元昊" in fn: return "元昊"
    if "邮政" in fn or "ems" in sn_all or "e邮宝" in sn_all: return "中国邮政"
    return "未知"

def channel_type(s):
    s=s.lower()
    if "空派" in s or "空运" in s or "空卡" in s: return "air"
    if "海派" in s or "海运" in s: return "sea"
    if "铁派" in s or "铁路" in s or "快铁" in s: return "rail"
    if "卡航" in s or "卡派" in s: return "truck"
    if "陆运" in s: return "land"
    if "快递" in s or "ups" in s or "dhl" in s or "fedex" in s: return "express"
    if "ems" in s or "e邮宝" in s or "e特快" in s or "小包" in s: return "postal"
    return "air"

def transport_type(s):
    s=s.lower()
    if "尾程" in s: return "仅海外尾程派送"
    if "干" in s: return "仅国际干线"
    if "双清" in s or "派送" in s or "到门" in s: return "干线+尾程双清派送"
    if "自提" in s or "到仓" in s: return "干线+尾程自提"
    return "干线+尾程双清派送"

def customs_mode(s):
    s=s.lower()
    if "ddp" in s or "包税" in s: return "DDP"
    if "ddu" in s or "dap" in s or "自税" in s: return "DDU/DAP"
    if "pva" in s or "递延" in s: return "PVA"
    if "自提" in s or "到仓" in s: return "双清到仓自提"
    if "cif" in s or "cfr" in s or "单清" in s: return "CIF/CFR"
    return "DDP"

def door_delivery(s):
    s=s.lower()
    if "到门" in s or "派送" in s: return True
    if "自提" in s or "到仓" in s or "到港" in s: return False
    return True

def make_id(supplier, product, country="", zone=""):
    sm = {"蓝牛国际":"LN","亿俐缇":"YLT","福鑫国际":"FX","韩润物流":"HR",
          "祥源国际":"XY","畅骏国际":"CJ","启丰物流":"QF","海源日本":"HY",
          "卉马国际":"HM","昌海运通":"CH","翔平":"XP","艾姆勒":"IML",
          "惠运捷邮政大包":"HYJ2","满天星综合":"MTX","满天星主营":"MTX2",
          "惠运捷":"HYJ","元昊":"YH","中国邮政":"CNPOST"}
    s = sm.get(supplier,"UNK")
    parts = [s]
    if product: parts.append(re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]','',product)[:12])
    if country: parts.append(country[:6])
    if zone: parts.append(re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]','',zone)[:6])
    return "-".join(parts)

def extract_weight_tiers(row):
    tiers = []
    for v in row:
        vs = ss(v).lower()
        m = re.match(r'^(\d+)\s*[-~]\s*(\d+)',vs)
        if m: tiers.append((int(m.group(1)),int(m.group(2))))
        else:
            m2 = re.match(r'^(\d+)\s*kg\+?',vs)
            if m2 and int(m2.group(1))>0: tiers.append((int(m2.group(1)),None))
    return tiers

def extract_prices(row, start=1, mx=50000):
    ps = []
    for v in row[start:]:
        pf = sf(v)
        if pf is not None and 0<pf<mx: ps.append(pf)
    return ps

def read_rows(ws, mr=100, mc=25):
    rows = []
    for ri in range(1, min(mr, ws.max_row+1)):
        rd = [cell(ws,ri,ci) for ci in range(1, min(mc, ws.max_column+1))]
        rows.append(rd)
    return rows

def extract_country(sn):
    m = {"美国":"美国","加拿大":"加拿大","英国":"英国","欧洲":"欧洲","澳大利亚":"澳大利亚",
         "新西兰":"新西兰","日本":"日本","韩国":"韩国","东南亚":"东南亚","俄罗斯":"俄罗斯",
         "南非":"南非","中东":"中东","南美":"南美","非洲":"非洲","台湾":"台湾","香港":"香港",
         "越南":"越南","泰国":"泰国","马来西亚":"马来西亚","新加坡":"新加坡","菲律宾":"菲律宾",
         "墨西哥":"墨西哥","巴西":"巴西","智利":"智利","哥伦比亚":"哥伦比亚","秘鲁":"秘鲁",
         "阿联酋":"阿联酋","沙特":"沙特","科威特":"科威特","卡塔尔":"卡塔尔","巴林":"巴林",
         "伊朗":"伊朗","阿曼":"阿曼","以色列":"以色列","约旦":"约旦","波兰":"波兰",
         "德国":"德国","法国":"法国","意大利":"意大利","西班牙":"西班牙","荷兰":"荷兰",
         "比利时":"比利时","卢森堡":"卢森堡","奥地利":"奥地利","克罗地亚":"克罗地亚",
         "匈牙利":"匈牙利","斯洛伐克":"斯洛伐克","斯洛文尼亚":"斯洛文尼亚","捷克":"捷克",
         "罗马尼亚":"罗马尼亚","保加利亚":"保加利亚","葡萄牙":"葡萄牙","瑞典":"瑞典",
         "芬兰":"芬兰","丹麦":"丹麦","挪威":"挪威","爱尔兰":"爱尔兰"}
    for k,v in m.items():
        if k in sn: return v
    return ""

def extract_delivery_days(text):
    m = re.search(r'(\d+)\s*[-~]\s*(\d+)\s*(?:个)?(?:工作)?天',text)
    if m: return int(m.group(1)), int(m.group(2))
    m = re.search(r'(\d+)\s*(?:个)?(?:工作)?天',text)
    if m: d=int(m.group(1)); return d,d
    return None,None

def extract_notes(row, keywords):
    notes = []
    for v in row:
        vs = ss(v)
        if vs and any(k in vs for k in keywords):
            notes.append(vs)
    return "; ".join(notes[:5])

def new_route(supplier, product, country, zone, ct, tt, **kw):
    rid = make_id(supplier, product, country, zone)
    postal = kw.get("postal_code_range","")
    sum_dim = kw.get("sum_dimension_limit")
    longest = kw.get("longest_side_limit")
    return {
        "id": rid, "provider_name": supplier, "name": product[:50],
        "cover_countries": country, "cover_cities": kw.get("cover_cities","全境"),
        "postal_code_range": postal if postal else "全境",
        "supported_categories": kw.get("supported_categories","普货"),
        "min_delivery_days": kw.get("min_delivery_days"),
        "max_delivery_days": kw.get("max_delivery_days"),
        "is_active": True,
        "packaging_restriction": kw.get("packaging_restriction","纸箱"),
        "channel_type": ct, "transport_type": tt,
        "door_delivery": kw.get("door_delivery",True),
        "sum_dimension_limit": sum_dim if sum_dim else "无",
        "longest_side_limit": longest if longest else "无",
        "customs_mode": kw.get("customs_mode","DDP"),
        "currency": kw.get("currency","CNY"),
        "remark": kw.get("remark",""),
    }

def new_pricing(route_name, **kw):
    postal = kw.get("applicable_postal_range","")
    region = kw.get("applicable_region","")
    return {
        "route_name": route_name, "countries": kw.get("countries",""),
        "applicable_region": region if region else "全境",
        "applicable_postal_range": postal if postal else "全境",
        "min_weight": kw.get("min_weight"), "max_weight": kw.get("max_weight"),
        "max_length": kw.get("max_length"), "min_length": kw.get("min_length"),
        "unit_price": kw.get("unit_price"), "unit_price_unit": kw.get("unit_price_unit","KG"),
        "first_weight_price": kw.get("first_weight_price"),
        "first_weight_unit": kw.get("first_weight_unit","KG"),
        "continued_weight_price": kw.get("continued_weight_price"),
        "continued_weight_unit": kw.get("continued_weight_unit","KG"),
        "min_charge": kw.get("min_charge"),
        "packaging_billing_unit": kw.get("packaging_billing_unit",""),
        "packaging_fee_wooden_crate": kw.get("packaging_fee_wooden_crate"),
        "packaging_fee_wooden_frame": kw.get("packaging_fee_wooden_frame"),
        "overlength_surcharge": kw.get("overlength_surcharge"),
        "overweight_surcharge": kw.get("overweight_surcharge"),
        "overgirth_surcharge": kw.get("overgirth_surcharge"),
        "tax_mode": kw.get("tax_mode",""),
        "registration_fee": kw.get("registration_fee"),
        "fuel_surcharge_rate": kw.get("fuel_surcharge_rate"),
        "customs_declaration_fee": kw.get("customs_declaration_fee"),
        "remote_area_surcharge": kw.get("remote_area_surcharge"),
        "pod_fee": kw.get("pod_fee"),
        "remark": kw.get("remark",""),
    }

def init_result(supplier, fname):
    return {"supplier_name":supplier,"file_name":fname,
            "parse_time":datetime.now().isoformat(),
            "routes":[],"pricing":[],"surcharges":[],"packaging_fees":[],"registration_fees":[]}

# ============================================================
# 通用重量段解析器
# ============================================================

def parse_weight_tier_sheet(ws, sheet_name, supplier, ct, tt, result, 
                            header_row=1, data_start=2, country_col=1, 
                            price_start_col=2, notes_col=None, max_rows=200):
    """通用重量段定价解析器"""
    rows = read_rows(ws, max_rows, 25)
    if len(rows) <= header_row: return
    
    # 提取重量段
    header = rows[header_row-1] if header_row > 0 else rows[0]
    tiers = extract_weight_tiers(header[price_start_col-1:])
    if not tiers: return
    
    current_product = sheet_name
    for i in range(data_start-1, len(rows)):
        row = rows[i]
        if row is None or all(v is None for v in row): continue
        country = ss(row[country_col-1]) if country_col <= len(row) else ""
        if not country: continue
        if re.match(r'^[一二三四五六七八九十]+[、．.]',country): 
            current_product = country; continue
        if any(k in country for k in ["附加费","赔偿","注意","说明","仓库代码"]): break
        
        prices = extract_prices(row, price_start_col-1)
        if not prices: continue
        
        # 收集备注
        all_text = " ".join([ss(v) for v in row if v])
        notes = extract_notes(row, ["限重","报关","计泡","DDU","DDP","关税","品名","限制",
                                     "禁运","电池","液体","粉末","带电","包税","自税","时效"])
        min_d, max_d = extract_delivery_days(all_text)
        cm = customs_mode(all_text)
        
        c = extract_country(sheet_name) or country
        route = new_route(supplier, f"{current_product[:20]}-{c}", c, "", ct, tt,
                         min_delivery_days=min_d, max_delivery_days=max_d,
                         customs_mode=cm, remark=notes[:300])
        result["routes"].append(route)
        
        for idx,(tmin,tmax) in enumerate(tiers):
            price = prices[idx] if idx < len(prices) else None
            if price is not None:
                result["pricing"].append(new_pricing(route["name"],
                    countries=c, min_weight=tmin, max_weight=tmax,
                    unit_price=price, unit_price_unit="KG",
                    remark=notes[:200]))


# ============================================================
# 蓝牛国际 - 美国超大件海运，按邮编分区
# ============================================================

def parse_lanniu(filepath, wb):
    result = init_result("蓝牛国际", os.path.basename(filepath))
    # 优先使用数据库sheet
    if "数据库" in wb.sheetnames:
        ws = wb["数据库"]
        rows = read_rows(ws, 100, 21)
        # R1: 服务区域|渠道代码|物流产品|目的港|邮编|23KG+|41KG+|...|3001KG+|派送费|参考时效
        header = rows[0]
        tiers = extract_weight_tiers(header[5:17])  # cols 6-17
        for row in rows[1:]:
            if row is None: continue
            zone = ss(row[0]); code = ss(row[1]); product = ss(row[2])
            port = ss(row[3]); postal = ss(row[4])
            if not product: continue
            prices = extract_prices(row, 5, 500)  # from col 6
            delivery_fee = sf(row[17]) if len(row) > 17 else None
            delivery_time = ss(row[18]) if len(row) > 18 else ""
            
            route = new_route("蓝牛国际", product[:30], "美国", zone[:15], "sea",
                            "干线+尾程双清派送",
                            postal_code_range=postal[:50],
                            cover_cities=port,
                            door_delivery=True, customs_mode="DDP",
                            remark=f"渠道:{code}; {delivery_time}")
            result["routes"].append(route)
            
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries="美国", applicable_region=zone[:15],
                        applicable_postal_range=postal[:30],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG",
                        min_charge=delivery_fee,
                        remark=f"派送费:{delivery_fee}元/票"))
    
    # 其他渠道sheet
    skip = {"目录","包装要求","风险告知书","赔付标准","报价快速查询","数据库",
            "增值服务","地址附加费","产品附加费","配载图","反倾销清单"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 100, 17)
        # 找header行(含重量段)
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG+" in ss(v) or "kg+" in ss(v) for v in row):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[3:])  # prices start from col 4
        
        current_product = sn
        current_port = ""
        current_code = ""
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            # 检查是否有新的产品/目的港
            b = ss(row[1]) if len(row)>1 else ""
            if b and ("直送" in b or "US" in b.upper() or "EC" in b.upper()):
                parts = b.split("\n")
                current_port = parts[0][:10]
                if len(parts)>1: current_code = parts[1][:10]
            
            zone = ss(row[2]) if len(row)>2 else ""
            if not zone or "邮编" not in zone: continue
            
            prices = extract_prices(row, 3, 500)
            if not prices: continue
            
            delivery_fee = sf(row[10]) if len(row)>10 else None
            delivery_time = ss(row[11]) if len(row)>11 else ""
            
            route = new_route("蓝牛国际", f"{current_product[:20]}-{current_port}", 
                            "美国", zone[:15], "sea", "干线+尾程双清派送",
                            postal_code_range=zone[:50],
                            cover_cities=current_port,
                            door_delivery=True, customs_mode="DDP",
                            remark=delivery_time)
            result["routes"].append(route)
            
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries="美国", applicable_region=zone[:15],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG",
                        min_charge=delivery_fee))
    return result


# ============================================================
# 亿俐缇 - 中东线路
# ============================================================

def parse_yiliti(filepath, wb):
    result = init_result("亿俐缇", os.path.basename(filepath))
    skip = {"报价目录"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 250, 20)
        
        # 检测格式：中东小包(按重量定价) vs 标准重量段
        if "小包" in sn or "T货" in sn:
            # 小包格式：行=重量(1KG,2KG...)，列=国家
            _parse_yiliti_small_packet(ws, sn, result)
            continue
        
        if "DHL" in sn.upper():
            _parse_yiliti_dhl(ws, sn, result)
            continue
        
        if "派送费" in sn or "海外仓" in sn:
            continue
        
        # 标准重量段格式
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        # 产品信息
        product_info = ss(rows[0][0]) if rows else sn
        ct = channel_type(sn); tt = transport_type(sn)
        country = extract_country(sn)
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c: continue
            if "渠道" in c or "编码" in c or "返回目录" in c: continue
            
            prices = extract_prices(row, 1)
            if not prices: continue
            
            all_text = " ".join([ss(v) for v in row])
            min_d, max_d = extract_delivery_days(all_text)
            
            route = new_route("亿俐缇", f"{sn[:15]}-{c[:8]}", country or c, "", ct, tt,
                            min_delivery_days=min_d, max_delivery_days=max_d,
                            door_delivery=True, customs_mode=customs_mode(all_text))
            result["routes"].append(route)
            
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=country or c, min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result

def _parse_yiliti_small_packet(ws, sn, result):
    rows = read_rows(ws, 40, 11)
    # R3: 国家|巴林|卡塔尔|科威特|沙特|时效
    # R6+: 1KG|530|630|560|910|时效
    if len(rows) < 6: return
    header = rows[2]  # 0-indexed
    countries = [ss(header[i]) for i in range(1, min(7, len(header))) if ss(header[i]) and "时效" not in ss(header[i])]
    
    for i in range(5, len(rows)):
        row = rows[i]
        if row is None: continue
        weight_str = ss(row[0])
        m = re.match(r'(\d+)\s*KG', weight_str)
        if not m: continue
        weight = int(m.group(1))
        
        for j, c in enumerate(countries):
            price = sf(row[j+1]) if j+1 < len(row) else None
            if price is None: continue
            
            route_name = f"中东T货小包-{c}"
            # 检查是否已有该路线
            existing = [r for r in result["routes"] if r["name"] == route_name]
            if not existing:
                country_full = extract_country(c) or c
                result["routes"].append(new_route("亿俐缇", route_name, country_full, "",
                    "air", "干线+尾程双清派送",
                    door_delivery=True, customs_mode="DDP",
                    supported_categories="特货",
                    remark="T货敏感均可运输; 单个包裹8KG以内"))
            
            result["pricing"].append(new_pricing(route_name,
                countries=extract_country(c) or c,
                min_weight=weight-1 if weight>1 else 0, max_weight=weight,
                unit_price=price/weight if weight>0 else price,
                unit_price_unit="KG",
                remark=f"{weight}KG整票价格:{price}元"))


def _parse_yiliti_dhl(ws, sn, result):
    rows = read_rows(ws, 150, 6)
    # 格式: 重量KG|价格
    for i, row in enumerate(rows):
        if row is None: continue
        weight_str = ss(row[0])
        m = re.match(r'([\d.]+)\s*(?:KG|kg)', weight_str)
        if not m: continue
        weight = float(m.group(1))
        price = sf(row[1])
        if price is None: continue
        
        route_name = f"DHL-{sn}"
        existing = [r for r in result["routes"] if r["name"] == route_name]
        if not existing:
            result["routes"].append(new_route("亿俐缇", route_name, "科威特", "",
                "express", "干线+尾程双清派送",
                door_delivery=True, customs_mode="DDP"))
        
        result["pricing"].append(new_pricing(route_name,
            countries="科威特",
            min_weight=weight, max_weight=weight,
            unit_price=price, unit_price_unit="KG",
            remark=f"{weight}KG整票价格"))


# ============================================================
# 福鑫国际 - 欧洲为主，分区+邮编+重量段
# ============================================================

def parse_fuxin(filepath, wb):
    result = init_result("福鑫国际", os.path.basename(filepath))
    skip = {"总目录","超大件包装要求","仓库地址和联系方式","产品附加费",
            "GLS偏远地区列表","欧洲海运超大件优先服务偏远邮编","东欧DPD偏远邮编"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 150, 30)
        
        ct = channel_type(sn); tt = transport_type(sn)
        
        # 找header行
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG+" in ss(v) for v in row):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        # 确定列位置
        zone_col = 0; city_col = 1; postal_col = 2; price_start = 3
        for ci, v in enumerate(header):
            vs = ss(v)
            if "分区" in vs: zone_col = ci
            elif "城市" in vs: city_col = ci
            elif "邮编" in vs: postal_col = ci
            elif "国家" in vs: zone_col = ci
        
        tiers = extract_weight_tiers(header[price_start:])
        if not tiers: continue
        
        # 产品名
        product = ss(rows[1][1]) if len(rows)>1 and len(rows[1])>1 else sn
        product = re.sub(r'福鑫[-—]?', '', product)[:25]
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            zone = ss(row[zone_col]) if zone_col < len(row) else ""
            city = ss(row[city_col]) if city_col < len(row) else ""
            postal = ss(row[postal_col]) if postal_col < len(row) else ""
            
            if not zone and not city: continue
            if "渠道说明" in zone or "下单" in zone: continue
            
            prices = extract_prices(row, price_start)
            if not prices: continue
            
            # 托盘费
            pallet_fee = None
            for ci in range(price_start, len(row)):
                if "托盘" in ss(header[ci]) if ci < len(header) else False:
                    pallet_fee = sf(row[ci]); break
            
            country = extract_country(sn) or zone
            route = new_route("福鑫国际", f"{product}-{country}", country, zone[:8],
                            ct, tt,
                            postal_code_range=postal[:50] if postal else "",
                            cover_cities=city[:20] if city else "",
                            door_delivery=True, customs_mode=customs_mode(sn),
                            remark=f"托盘费:{pallet_fee}" if pallet_fee else "")
            result["routes"].append(route)
            
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=country, applicable_region=zone[:10],
                        applicable_postal_range=postal[:30],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG",
                        packaging_billing_unit="托盘" if pallet_fee else "",
                        remark=f"托盘费:{pallet_fee}元" if pallet_fee else ""))
    return result


# ============================================================
# 韩润物流 - 韩国专线
# ============================================================

def parse_hanrun(filepath, wb):
    result = init_result("韩润物流", os.path.basename(filepath))
    skip = {"目录","注意事项","禁运清单","仓库地址"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 100, 15)
        ct = channel_type(sn); tt = transport_type(sn)
        
        # 找header
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c or "备注" in c: continue
            prices = extract_prices(row, 1)
            if not prices: continue
            
            route = new_route("韩润物流", f"{sn[:15]}-{c[:8]}", "韩国", "", ct, tt,
                            door_delivery=True, customs_mode="DDP")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries="韩国", min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 祥源国际 - 南美专线
# ============================================================

def parse_xiangyuan(filepath, wb):
    result = init_result("祥源国际", os.path.basename(filepath))
    skip = {"首页","目录","注意事项","仓库地址","箱单发票模板"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 50, 15)
        ct = channel_type(sn)
        country = extract_country(sn)
        
        # 检测格式：CBM计价 vs KG计价
        is_cbm = False
        for row in rows:
            if row and any("CBM" in ss(v) for v in row):
                is_cbm = True; break
        
        if is_cbm:
            _parse_xiangyuan_cbm(ws, sn, country, ct, result)
        else:
            _parse_xiangyuan_kg(ws, sn, country, ct, result)
    return result

def _parse_xiangyuan_cbm(ws, sn, country, ct, result):
    """CBM计价：产品分类在列，重量段(CBM段)在行"""
    rows = read_rows(ws, 50, 15)
    # R3: 产品分类 | 品类1 | ... | 品类2 | ...
    # R4: 1-5CBM | RMB3400/CBM | ... | RMB3600/CBM | ...
    # R6: 10CBM以上货物 | RMB3300/CBM | ...
    
    # 提取品类列
    header = rows[2] if len(rows) > 2 else []
    categories = []
    for ci in range(1, len(header)):
        cat = ss(header[ci])
        if cat: categories.append((ci, cat[:30]))
    if not categories: categories = [(1, "普货")]
    
    # 提取CBM段
    cbm_tiers = []
    for i in range(3, len(rows)):
        row = rows[i]
        if row is None: continue
        vol = ss(row[0])
        if not vol: continue
        m = re.match(r'(\d+)\s*[-~]\s*(\d+)\s*CBM', vol)
        if m:
            cbm_tiers.append((i, f"{m.group(1)}-{m.group(2)}CBM", int(m.group(1)), int(m.group(2))))
        else:
            m2 = re.match(r'(\d+)\s*CBM', vol)
            if m2:
                cbm_tiers.append((i, f"{m2.group(1)}CBM+", int(m2.group(1)), None))
    
    if not cbm_tiers: return
    
    for row_idx, tier_name, min_v, max_v in cbm_tiers:
        row = rows[row_idx]
        for ci, cat_name in categories:
            price_str = ss(row[ci]) if ci < len(row) else ""
            m = re.search(r'(\d+)', price_str)
            if not m: continue
            price = int(m.group(1))
            
            route = new_route("祥源国际", f"{sn[:15]}-{cat_name[:10]}", country, "",
                            ct, "干线+尾程双清派送",
                            door_delivery=False, customs_mode="DDP",
                            supported_categories=cat_name[:20],
                            remark=f"CBM段:{tier_name}; 双清包税到仓")
            result["routes"].append(route)
            result["pricing"].append(new_pricing(route["name"],
                countries=country,
                min_weight=min_v, max_weight=max_v,
                unit_price=price, unit_price_unit="CBM",
                remark=f"{tier_name}: {price}元/CBM"))

def _parse_xiangyuan_kg(ws, sn, country, ct, result):
    """KG计价：品类在列，重量段在行"""
    rows = read_rows(ws, 50, 15)
    # R3: 产品分类 | 品类1 | 品类2 | ...
    # R8: 21KG+ | RMB95/KG | RMB108/KG | ...
    
    # 提取品类列
    header = None
    for i, row in enumerate(rows):
        if row and "产品分类" in ss(row[0]):
            header = i; break
    if header is None: return
    
    header_row = rows[header]
    categories = []
    for ci in range(1, len(header_row)):
        cat = ss(header_row[ci])
        if cat: categories.append((ci, cat[:30]))
    if not categories: categories = [(1, "普货")]
    
    # 提取重量段
    for i in range(header+1, len(rows)):
        row = rows[i]
        if row is None: continue
        weight_str = ss(row[0])
        m = re.match(r'(\d+)\s*KG', weight_str)
        if not m: continue
        min_w = int(m.group(1))
        
        for ci, cat_name in categories:
            price_str = ss(row[ci]) if ci < len(row) else ""
            m2 = re.search(r'(\d+)', price_str)
            if not m2: continue
            price = int(m2.group(1))
            
            route = new_route("祥源国际", f"{sn[:15]}-{cat_name[:10]}-{min_w}KG+", country, "",
                            ct, "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP",
                            supported_categories=cat_name[:20])
            result["routes"].append(route)
            result["pricing"].append(new_pricing(route["name"],
                countries=country, min_weight=min_w,
                unit_price=price, unit_price_unit="KG"))


# ============================================================
# 畅骏国际 - FBA仓库代码定价
# ============================================================

def parse_changjun(filepath, wb):
    result = init_result("畅骏国际", os.path.basename(filepath))
    skip = {"目录","导航","注意事项","仓库地址","反倾销"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 50, 15)
        ct = channel_type(sn)
        
        # 找header
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG+" in ss(v) for v in row):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        product = sn[:20]
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            warehouse = ss(row[0])
            if not warehouse or "仓库" in warehouse: continue
            
            prices = extract_prices(row, 1)
            if not prices: continue
            
            all_text = " ".join([ss(v) for v in row])
            min_d, max_d = extract_delivery_days(all_text)
            
            route = new_route("畅骏国际", f"{product}-{warehouse[:15]}", "美国", warehouse[:15],
                            ct, "干线+尾程双清派送",
                            min_delivery_days=min_d, max_delivery_days=max_d,
                            door_delivery=True, customs_mode="DDP")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries="美国", applicable_region=warehouse[:20],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 启丰物流 - 按地区(美西/美中/美东)+重量段
# ============================================================

def parse_qifeng(filepath, wb):
    result = init_result("启丰物流", os.path.basename(filepath))
    skip = {"启丰报价目录","空海派发货须知"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 60, 18)
        ct = channel_type(sn)
        
        # 找header行 - 动态检测重量段起始列
        header_idx = None
        price_offset = None
        for i, row in enumerate(rows):
            if row is None: continue
            for ci in range(len(row)):
                v = ss(row[ci])
                if re.search(r'\d+\s*[-~]?\s*\d*\s*KG', v, re.I):
                    header_idx = i
                    price_offset = ci
                    break
            if header_idx is not None: break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[price_offset:])
        if not tiers: continue
        
        product = sn[:25]
        country = extract_country(sn)
        
        # 确定region列(在price_offset之前)
        region_col = price_offset - 2 if price_offset >= 2 else 0
        channel_col = price_offset - 1 if price_offset >= 1 else 0
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            region = ss(row[region_col]) if region_col < len(row) else ""
            channel = ss(row[channel_col]) if channel_col < len(row) else ""
            
            if not region: continue
            if "返回" in region or "目录" in region or "目的地" in region: continue
            if re.match(r'^超\d', region): continue  # skip notes like "超3个工作日..."
            
            prices = extract_prices(row, price_offset)
            if not prices: continue
            
            all_text = " ".join([ss(v) for v in row])
            min_d, max_d = extract_delivery_days(all_text)
            
            route = new_route("启丰物流", f"{product}-{region[:8]}", country, region[:10],
                            ct, "干线+尾程双清派送",
                            min_delivery_days=min_d, max_delivery_days=max_d,
                            door_delivery=True, customs_mode=customs_mode(all_text),
                            remark=channel[:50] if channel else "")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=country, applicable_region=region[:10],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 海源日本 - 日本专线多渠道
# ============================================================

def parse_haiyuan(filepath, wb):
    result = init_result("海源日本", os.path.basename(filepath))
    skip = {"目录","注意事项","仓库地址"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 100, 15)
        ct = channel_type(sn)
        
        # 找header
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c or "备注" in c: continue
            prices = extract_prices(row, 1)
            if not prices: continue
            
            route = new_route("海源日本", f"{sn[:15]}-{c[:8]}", "日本", "", ct,
                            "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries="日本", min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 卉马国际 - 欧洲/美国/英国
# ============================================================

def parse_huima(filepath, wb):
    result = init_result("卉马国际", os.path.basename(filepath))
    skip = {"目录","注意事项","仓库地址","免责声明"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 100, 20)
        ct = channel_type(sn)
        
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        country = extract_country(sn)
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c or "备注" in c or "说明" in c: continue
            prices = extract_prices(row, 1)
            if not prices: continue
            
            all_text = " ".join([ss(v) for v in row])
            min_d, max_d = extract_delivery_days(all_text)
            
            dest = country or extract_country(c) or c
            route = new_route("卉马国际", f"{sn[:15]}-{dest}", dest, "", ct,
                            "干线+尾程双清派送",
                            min_delivery_days=min_d, max_delivery_days=max_d,
                            door_delivery=True, customs_mode=customs_mode(all_text))
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=dest, min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 昌海运通 - 墨西哥专线
# ============================================================

def parse_changhai(filepath, wb):
    result = init_result("昌海运通", os.path.basename(filepath))
    # .xls format
    if filepath.endswith('.xls'):
        if xlrd is None:
            print(json.dumps({"error":"需要xlrd库解析.xls文件"},ensure_ascii=False))
            return result
        xls_wb = xlrd.open_workbook(filepath)
        for sn in xls_wb.sheet_names():
            ws = xls_wb.sheet_by_name(sn)
            if ws.nrows < 5: continue
            # 找header
            header_row = None
            for r in range(ws.nrows):
                vals = [str(ws.cell_value(r,c)) for c in range(ws.ncols)]
                if any("KG" in v.upper() or "单价" in v for v in vals):
                    header_row = r; break
            if header_row is None: continue
            
            # 解析数据行
            for r in range(header_row+1, ws.nrows):
                cat = str(ws.cell_value(r, 0)).strip()
                if not cat or "备注" in cat or "类" not in cat: continue
                code = str(ws.cell_value(r, 1)).strip()
                price_str = str(ws.cell_value(r, 2)).strip()
                m = re.search(r'(\d+)\s*/\s*KG', price_str)
                price = int(m.group(1)) if m else sf(price_str)
                if price is None: continue
                
                # 提取重量门槛
                m2 = re.search(r'\((\d+)KG\+\)', price_str)
                min_w = int(m2.group(1)) if m2 else 10
                
                route = new_route("昌海运通", f"墨西哥海派-{cat[:8]}", "墨西哥", "",
                                "sea", "干线+尾程双清派送",
                                door_delivery=True, customs_mode="DDP",
                                supported_categories=cat,
                                remark=code)
                result["routes"].append(route)
                result["pricing"].append(new_pricing(route["name"],
                    countries="墨西哥", min_weight=min_w,
                    unit_price=price, unit_price_unit="KG"))
        return result
    
    # .xlsx format
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = read_rows(ws, 50, 10)
        for i, row in enumerate(rows):
            if row is None: continue
            cat = ss(row[0])
            if not cat or "类" not in cat: continue
            price_str = ss(row[2]) if len(row)>2 else ""
            m = re.search(r'(\d+)\s*/\s*KG', price_str)
            if not m: continue
            price = int(m.group(1))
            m2 = re.search(r'\((\d+)KG\+\)', price_str)
            min_w = int(m2.group(1)) if m2 else 10
            
            route = new_route("昌海运通", f"墨西哥海派-{cat[:8]}", "墨西哥", "",
                            "sea", "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP",
                            supported_categories=cat)
            result["routes"].append(route)
            result["pricing"].append(new_pricing(route["name"],
                countries="墨西哥", min_weight=min_w,
                unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 翔平
# ============================================================

def parse_xiangping(filepath, wb):
    result = init_result("翔平", os.path.basename(filepath))
    skip = {"目录","注意事项","仓库地址"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 100, 15)
        ct = channel_type(sn)
        country = extract_country(sn)
        
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c or "备注" in c: continue
            prices = extract_prices(row, 1)
            if not prices: continue
            
            dest = country or extract_country(c) or c
            route = new_route("翔平", f"{sn[:15]}-{dest}", dest, "", ct,
                            "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=dest, min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 艾姆勒 - 俄罗斯专线
# ============================================================

def parse_iml(filepath, wb):
    result = init_result("艾姆勒", os.path.basename(filepath))
    skip = {"目录","注意事项","仓库地址","免责声明"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws, 50, 15)
        
        # 艾姆勒格式：类型|重量KG|运费RMB/KG|挂号费|保险|备注
        for i, row in enumerate(rows):
            if row is None: continue
            product_type = ss(row[0])
            if not product_type or "IML" not in product_type.upper(): continue
            if "停收" in product_type: continue
            
            weight_str = ss(row[1]) if len(row)>1 else ""
            price = sf(row[2]) if len(row)>2 else None
            reg_fee = sf(row[4]) if len(row)>4 else None
            notes = ss(row[7]) if len(row)>7 else ""
            
            if price is None: continue
            
            # 解析重量范围
            m = re.match(r'([\d.]+)\s*[-~]\s*([\d.]+)\s*KG', weight_str)
            min_w = float(m.group(1)) if m else 0
            max_w = float(m.group(2)) if m else 30
            
            ct = "land" if "陆运" in product_type else "express" if "快递" in product_type else "air"
            
            cats = "普货"
            if "服装" in product_type or "鞋" in product_type: cats = "特货"
            if "带电" in product_type: cats = "带电"
            
            route = new_route("艾姆勒", f"{sn[:12]}-{product_type[:12]}", "俄罗斯", "",
                            ct, "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP",
                            supported_categories=cats, remark=notes[:200])
            result["routes"].append(route)
            result["pricing"].append(new_pricing(route["name"],
                countries="俄罗斯", min_weight=min_w, max_weight=max_w,
                unit_price=price, unit_price_unit="KG",
                registration_fee=reg_fee,
                remark=notes[:200]))
    return result


# ============================================================
# 惠运捷邮政大包
# ============================================================

def parse_hyj_postal_big(filepath, wb):
    result = init_result("惠运捷邮政大包", os.path.basename(filepath))
    if "邮政大包" not in wb.sheetnames:
        return result
    ws = wb["邮政大包"]
    rows = read_rows(ws, 250, 18)
    
    # R7: 首/续重1千克 | 航空 | | 空运水陆路(SAL) | | 水陆路 | | 限重 | 备注
    # R8+: 国家 | 首重 | 续重 | 首重 | 续重 | 首重 | 续重 | 限重 | | 备注
    for i in range(7, len(rows)):
        row = rows[i]
        if row is None: continue
        country = ss(row[0])
        if not country or "注意" in country or "通达" in country: continue
        
        af = sf(row[1]); ac = sf(row[2])  # 航空
        sf2 = sf(row[3]); sc = sf(row[4])  # SAL
        wf = sf(row[5]); wc = sf(row[6])   # 水陆路
        wl = sf(row[7])  # 限重
        notes = ss(row[9]) if len(row)>9 else ""
        
        for tp, fp, cp in [("航空",af,ac),("SAL",sf2,sc),("水陆路",wf,wc)]:
            if fp is None and cp is None: continue
            name = f"邮政大包-{tp}-{country}"
            result["routes"].append(new_route("惠运捷邮政大包", name, country, "",
                "postal", "干线+尾程双清派送",
                door_delivery=True, customs_mode="DDP",
                remark=f"限重{int(wl)}kg; {notes}" if wl else notes))
            result["pricing"].append(new_pricing(name,
                countries=country,
                first_weight_price=fp, first_weight_unit="KG",
                continued_weight_price=cp, continued_weight_unit="KG",
                min_weight=1, max_weight=wl,
                customs_declaration_fee=8,
                remark="首重1kg+续重1kg; 报关费8元/票"))
    return result


# ============================================================
# 启丰报价表(.xls)
# ============================================================

def parse_qifeng_xls(filepath):
    result = init_result("启丰物流", os.path.basename(filepath))
    if xlrd is None:
        print(json.dumps({"error":"需要xlrd库解析.xls文件"},ensure_ascii=False))
        return result
    
    xls_wb = xlrd.open_workbook(filepath)
    skip = {"交仓地址"}
    for sn in xls_wb.sheet_names():
        if sn in skip: continue
        ws = xls_wb.sheet_by_name(sn)
        if ws.nrows < 4: continue
        ct = channel_type(sn)
        country = extract_country(sn)
        
        # 找header行并检测重量段起始列
        header_row = None
        price_col_start = None
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                v = str(ws.cell_value(r,c)).strip()
                if re.search(r'\d+\s*KG\+?', v, re.I):
                    header_row = r
                    price_col_start = c
                    break
            if header_row is not None: break
        if header_row is None: continue
        
        header_vals = [str(ws.cell_value(header_row,c)) for c in range(ws.ncols)]
        tiers = extract_weight_tiers(header_vals[price_col_start:])
        if not tiers: continue
        
        # 确定region列
        region_col = price_col_start - 1 if price_col_start >= 1 else 0
        channel_col = price_col_start - 2 if price_col_start >= 2 else -1
        
        product = sn[:20]
        current_channel = ""
        for r in range(header_row+1, ws.nrows):
            # 检查是否有新的渠道名
            if channel_col >= 0:
                ch = str(ws.cell_value(r, channel_col)).strip()
                if ch and ch != "服务渠道":
                    current_channel = ch
            
            region = str(ws.cell_value(r, region_col)).strip()
            if not region or "国家" in region or "服务渠道" in region: continue
            
            prices = []
            for c in range(price_col_start, min(ws.ncols, price_col_start+len(tiers))):
                v = ws.cell_value(r, c)
                pf = sf(v)
                if pf is not None and pf > 0: prices.append(pf)
            if not prices: continue
            
            route = new_route("启丰物流", f"{product}-{region[:12]}", country, region[:15],
                            ct, "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP",
                            remark=current_channel[:50])
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=country, applicable_region=region[:15],
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result


# ============================================================
# 原有5个货代解析器（保持不变）
# ============================================================

def parse_manTianXing_comprehensive(filepath, wb):
    result = init_result("满天星", os.path.basename(filepath))
    skip = {"目录","出货须知","免责声明与赔偿条款","俄罗斯制裁HScode","澳洲-常见海外仓地址"}
    for sn in wb.sheetnames:
        if sn in skip or "船期" in sn: continue
        ws = wb[sn]
        if "附加费" in sn or "仓库增值服务" in sn or "杂费" in sn: continue
        rows = read_rows(ws,50,20)
        ct = channel_type(sn); tt = transport_type(sn)
        current_product = ""
        for row in rows:
            if row and row[0]:
                rs = ss(row[0])
                if len(rs)>3 and not re.match(r'^[\d\.]+$',rs):
                    current_product = rs; break
        if not current_product: current_product = sn
        for i, row in enumerate(rows):
            wt = extract_weight_tiers(row)
            if not wt: continue
            for j in range(i+1, min(i+30, len(rows))):
                dr = rows[j]
                if dr is None or all(v is None for v in dr): continue
                zn = ss(dr[0])
                if not zn: continue
                if re.match(r'^[一二三四五六七八九十]+[、．.]',zn): break
                if any(k in zn for k in ["附加费","赔偿","注意","说明","仓库代码"]): break
                prices = extract_prices(dr, 1)
                if not prices: continue
                country = extract_country(sn)
                route = new_route("满天星", current_product[:30], country, zn[:10], ct, tt,
                                door_delivery=True, customs_mode="DDP")
                result["routes"].append(route)
                for idx,(tmin,tmax) in enumerate(wt):
                    price = prices[idx] if idx < len(prices) else None
                    if price is not None:
                        result["pricing"].append(new_pricing(route["name"],
                            countries=country, min_weight=tmin, max_weight=tmax,
                            unit_price=price, unit_price_unit="KG"))
    return result

def parse_manTianXing_main(filepath, wb):
    result = init_result("满天星", os.path.basename(filepath))
    skip = {"目录","出货须知","反倾销名录","海派偏远查询","加拿大仓库地址","美国仓库地址",
            "邮编库","加拿大FSA","FBA卡派速查","映射表","美国偏远邮编","美国禁运岛屿邮编",
            "加拿大偏远邮编","美加尾端卡车测算"}
    for sn in wb.sheetnames:
        if sn in skip or "船期" in sn: continue
        ws = wb[sn]
        if "仓库增值服务" in sn or "尾程价格表" in sn or "FBA价格表" in sn: continue
        rows = read_rows(ws,60,20)
        ct = channel_type(sn); tt = transport_type(sn)
        current_product = ""
        for row in rows:
            if row and row[0]:
                rs = ss(row[0])
                if len(rs)>3 and not re.match(r'^[\d\.]+$',rs):
                    current_product = rs; break
        if not current_product: current_product = sn
        for i, row in enumerate(rows):
            wt = extract_weight_tiers(row)
            if not wt: continue
            for j in range(i+1, min(i+30, len(rows))):
                dr = rows[j]
                if dr is None or all(v is None for v in dr): continue
                zn = ss(dr[0])
                if not zn: continue
                if re.match(r'^[一二三四五六七八九十]+[、．.]',zn): break
                if any(k in zn for k in ["附加费","赔偿","注意","说明","仓库代码"]): break
                prices = extract_prices(dr, 1)
                if not prices: continue
                country = extract_country(sn)
                route = new_route("满天星", current_product[:30], country, zn[:10], ct, tt,
                                door_delivery=True, customs_mode="DDP")
                result["routes"].append(route)
                for idx,(tmin,tmax) in enumerate(wt):
                    price = prices[idx] if idx < len(prices) else None
                    if price is not None:
                        result["pricing"].append(new_pricing(route["name"],
                            countries=country, min_weight=tmin, max_weight=tmax,
                            unit_price=price, unit_price_unit="KG"))
    return result

def parse_huiYunJie(filepath, wb):
    result = init_result("惠运捷", os.path.basename(filepath))
    skip = {"注意事项","2023邮政境外海关电话","国外邮政通讯录","邮政跟踪网址"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        rows = read_rows(ws,50,15)
        ct = channel_type(sn)
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("KG" in ss(v).upper() for v in row[1:]):
                header_idx = i; break
        if header_idx is None: continue
        header = rows[header_idx]
        tiers = extract_weight_tiers(header[1:])
        if not tiers: continue
        for i in range(header_idx+1, len(rows)):
            row = rows[i]
            if row is None: continue
            c = ss(row[0])
            if not c: continue
            prices = extract_prices(row, 1)
            if not prices: continue
            route = new_route("惠运捷", f"{sn[:15]}-{c[:8]}", extract_country(sn) or c, "",
                            ct, "干线+尾程双清派送",
                            door_delivery=True, customs_mode="DDP")
            result["routes"].append(route)
            for idx,(tmin,tmax) in enumerate(tiers):
                price = prices[idx] if idx < len(prices) else None
                if price is not None:
                    result["pricing"].append(new_pricing(route["name"],
                        countries=extract_country(sn) or c,
                        min_weight=tmin, max_weight=tmax,
                        unit_price=price, unit_price_unit="KG"))
    return result

def parse_yuanHao(filepath, wb):
    """元昊国际物流解析器 - 严格只提取路线信息"""
    result = init_result("元昊", os.path.basename(filepath))
    
    # 跳过非路线sheet
    skip_sheets = {"海外仓价格表", "FBS 价格表", "FBS价格表", "空派系统"}
    
    for sn in wb.sheetnames:
        if sn in skip_sheets:
            continue
        if "附加费" in sn:
            continue
            
        ws = wb[sn]
        rows = read_rows(ws, 100, 20)
        if len(rows) < 3:
            continue
        
        # 判断sheet类型并解析
        if ("俄罗斯" in sn or "白俄罗斯" in sn or "哈萨克" in sn) and "海外仓" not in sn and "FBS" not in sn:
            _parse_yh_main_route(ws, sn, result)
        elif "重货" in sn:
            _parse_yh_heavy_cargo(ws, sn, result)
        elif "单清" in sn:
            _parse_yh_single_customs(ws, sn, result)
    
    return result


def _parse_price_text(text):
    """解析价格文本，返回 (单价, 首重价, 续重价, 首重单位)"""
    text = ss(text)
    if not text or text == "/":
        return None, None, None, None
    
    # 首重+续重格式: "首重1KG80元+续重1KG35元"
    m = re.search(r'首重\s*(\d+\.?\d*)\s*KG?\s*(\d+\.?\d*)\s*元.*续重\s*\d+\.?\d*\s*KG?\s*(\d+\.?\d*)\s*元', text)
    if m:
        first_weight = float(m.group(1))
        first_price = float(m.group(2))
        cont_price = float(m.group(3))
        return None, first_price, cont_price, f"{first_weight}KG"
    
    # 单价格式: "35元/KG" 或 "35元/kg"
    m = re.search(r'(\d+\.?\d*)\s*元\s*/\s*KG', text, re.IGNORECASE)
    if m:
        return float(m.group(1)), None, None, "KG"
    
    # 纯数字
    pf = sf(text)
    if pf and pf > 0:
        return pf, None, None, "KG"
    
    return None, None, None, None


def _parse_yh_main_route(ws, sheet_name, result):
    """解析元昊主要路线sheet（俄罗斯/白俄罗斯/哈萨克）"""
    rows = read_rows(ws, 100, 20)
    
    # 找header行（含重量段如"0.1-15KG"）
    header_idx = None
    header = None
    for i, row in enumerate(rows):
        if row is None:
            continue
        row_text = " ".join([ss(v) for v in row])
        if "KG" in row_text.upper() and ("-" in row_text or "以上" in row_text):
            # 检查是否包含重量段格式
            if re.search(r'\d+\.?\d*\s*[-~]\s*\d+\.?\d*\s*KG', row_text, re.IGNORECASE):
                header_idx = i
                header = row
                break
    
    if header_idx is None or header is None:
        return
    
    # 提取重量段（从header行）
    tiers = []
    for v in header:
        vs = ss(v)
        # 匹配 "0.1-15KG", "15KG-40KG", "40KG以上"
        m = re.match(r'(\d+\.?\d*)\s*[-~]\s*(\d+\.?\d*)\s*KG', vs, re.IGNORECASE)
        if m:
            tiers.append((float(m.group(1)), float(m.group(2))))
        elif re.search(r'(\d+\.?\d*)\s*KG\s*以上', vs, re.IGNORECASE):
            m2 = re.search(r'(\d+\.?\d*)\s*KG\s*以上', vs, re.IGNORECASE)
            if m2:
                tiers.append((float(m2.group(1)), None))
    
    if not tiers:
        return
    
    # 确定价格列位置（从header行找）
    price_start_col = 4  # 默认从C5开始（index 4）
    for ci, v in enumerate(header):
        vs = ss(v)
        if re.search(r'\d+\.?\d*\s*[-~]\s*\d+\.?\d*\s*KG', vs, re.IGNORECASE):
            price_start_col = ci
            break
    
    # 遍历数据行
    current_route = None
    current_category = ""
    current_customs = ""
    
    for i in range(header_idx + 1, min(header_idx + 50, len(rows))):
        row = rows[i]
        if row is None:
            continue
        
        c1 = ss(row[0]) if len(row) > 0 else ""
        c2 = ss(row[1]) if len(row) > 1 else ""
        c3 = ss(row[2]) if len(row) > 2 else ""
        c4 = ss(row[3]) if len(row) > 3 else ""
        
        # 检查是否是新的路线（C1包含路线信息如"广州-莫斯科"）
        if c1 and ("-" in c1 or "时效" in c1) and len(c1) > 5:
            # 提取路线名称
            route_match = re.match(r'([^-]+-[^ ]+)', c1)
            if route_match:
                route_name = route_match.group(1).strip()
                current_route = route_name
                current_category = c2[:20] if c2 else ""
                current_customs = c3[:10] if c3 else ""
        
        # 检查是否是分区价格行（C4包含"一区"/"二区"/"三区"）
        if c4 and ("区" in c4) and current_route:
            zone = c4[:10]
            
            # 提取价格
            for ti, (tmin, tmax) in enumerate(tiers):
                col_idx = price_start_col + ti
                if col_idx >= len(row):
                    break
                
                price_text = ss(row[col_idx]) if row[col_idx] else ""
                unit_price, first_price, cont_price, unit = _parse_price_text(price_text)
                
                if unit_price or first_price:
                    # 确定国家
                    country = "俄罗斯"
                    if "白俄罗斯" in sheet_name:
                        country = "白俄罗斯"
                    elif "哈萨克" in sheet_name:
                        country = "哈萨克斯坦"
                    
                    # 创建路线（不含分区后缀，保持路线名称干净）
                    route = new_route("元昊", current_route, country, "",
                                    "land", "干线+尾程双清派送",
                                    door_delivery=True, customs_mode="DDP",
                                    supported_categories=current_category,
                                    remark=f"品类:{current_category}; 关税:{current_customs}")
                    
                    # 检查是否已存在该路线，避免重复
                    existing = [r for r in result["routes"] if r["name"] == current_route and r["cover_countries"] == country]
                    if not existing:
                        result["routes"].append(route)
                    
                    # 创建定价记录，分区信息放在applicable_region
                    pricing = new_pricing(current_route,
                        countries=country,
                        applicable_region=zone,  # 分区信息（一区/二区/三区）
                        min_weight=tmin, max_weight=tmax,
                        unit_price=unit_price, unit_price_unit="KG",
                        first_weight_price=first_price, first_weight_unit=unit or "KG",
                        continued_weight_price=cont_price, continued_weight_unit=unit or "KG",
                        remark=f"分区:{zone}")
                    result["pricing"].append(pricing)


def _parse_yh_heavy_cargo(ws, sheet_name, result):
    """解析元昊重货价格表"""
    rows = read_rows(ws, 50, 10)
    
    # 找header行
    header_idx = None
    tiers = []
    for i, row in enumerate(rows):
        if row is None:
            continue
        row_text = " ".join([ss(v) for v in row])
        if "重量区间" in row_text or "KG" in row_text.upper():
            header_idx = i
            # 提取重量段
            for v in row:
                vs = ss(v)
                m = re.match(r'(\d+)\s*[-~]\s*(\d+)\s*KG', vs, re.IGNORECASE)
                if m:
                    tiers.append((int(m.group(1)), int(m.group(2))))
                elif re.search(r'(\d+)\s*KG', vs, re.IGNORECASE) and "以上" in vs:
                    m2 = re.search(r'(\d+)\s*KG', vs, re.IGNORECASE)
                    if m2:
                        tiers.append((int(m2.group(1)), None))
            break
    
    if not tiers:
        return
    
    # 遍历数据行
    for i in range(header_idx + 1, min(header_idx + 20, len(rows))):
        row = rows[i]
        if row is None:
            continue
        
        route_text = ss(row[0]) if len(row) > 0 else ""
        if not route_text or "注意" in route_text or "禁止" in route_text:
            continue
        
        # 提取路线名称
        route_match = re.match(r'([^-]+-[^-]+)', route_text)
        if not route_match:
            continue
        
        route_name = route_match.group(1).strip()[:30]
        
        # 提取价格
        for ti, (tmin, tmax) in enumerate(tiers):
            col_idx = 2 + ti  # 价格从C3开始
            if col_idx >= len(row):
                break
            
            price_text = ss(row[col_idx]) if row[col_idx] else ""
            unit_price, _, _, _ = _parse_price_text(price_text)
            
            if unit_price:
                route = new_route("元昊", f"重货-{route_name}", "俄罗斯", "",
                                "land", "干线+尾程双清派送",
                                door_delivery=True, customs_mode="DDP",
                                supported_categories="重货/设备")
                result["routes"].append(route)
                
                pricing = new_pricing(f"重货-{route_name}",
                    countries="俄罗斯",
                    min_weight=tmin,
                    max_weight=tmax,
                    unit_price=unit_price,
                    unit_price_unit="KG")
                result["pricing"].append(pricing)


def _parse_yh_single_customs(ws, sheet_name, result):
    """解析元昊单清价格表"""
    rows = read_rows(ws, 50, 5)
    
    # 找header行
    header_idx = None
    for i, row in enumerate(rows):
        if row is None:
            continue
        row_text = " ".join([ss(v) for v in row])
        if "重量区间" in row_text:
            header_idx = i
            break
    
    if header_idx is None:
        return
    
    # 遍历数据行
    for i in range(header_idx + 1, min(header_idx + 15, len(rows))):
        row = rows[i]
        if row is None:
            continue
        
        weight_text = ss(row[0]) if len(row) > 0 else ""
        price_text = ss(row[1]) if len(row) > 1 else ""
        
        if not weight_text or not price_text:
            continue
        
        # 解析重量范围
        m = re.match(r'(\d+)\s*[-~]\s*(\d+)\s*(?:kg|以下|以上)?', weight_text, re.IGNORECASE)
        if not m:
            # 尝试 "100以下" 格式
            m = re.match(r'(\d+)\s*以下', weight_text)
            if m:
                tmin, tmax = 0, int(m.group(1))
            else:
                continue
        else:
            tmin, tmax = int(m.group(1)), int(m.group(2))
        
        # 解析价格
        unit_price, _, _, _ = _parse_price_text(price_text)
        if not unit_price:
            continue
        
        route_name = "单清-满洲里-俄罗斯"
        # 检查是否已存在该路线，避免重复
        existing = [r for r in result["routes"] if r["name"] == route_name]
        if not existing:
            route = new_route("元昊", route_name, "俄罗斯", "",
                            "land", "仅国际干线",
                            door_delivery=False, customs_mode="CIF/CFR",
                            remark="单清（不含进口清关）")
            result["routes"].append(route)
        
        pricing = new_pricing(route_name,
            countries="俄罗斯",
            min_weight=tmin,
            max_weight=tmax,
            unit_price=unit_price,
            unit_price_unit="KG",
            customs_declaration_fee=600,
            remark="中国报关费600元/票")
        result["pricing"].append(pricing)

def parse_chinaPost(filepath, wb):
    result = init_result("中国邮政", os.path.basename(filepath))
    skip = {"目 录","目录","禁限寄规定","常见问题","VAT政策","保价赔偿"}
    for sn in wb.sheetnames:
        if sn in skip: continue
        ws = wb[sn]
        if "EMS" in sn: _cp_ems(ws,result)
        elif "e特快" in sn: _cp_e_express(ws,result)
        elif "e邮宝" in sn: _cp_eub(ws,sn,result)
        elif "国际小包" in sn: _cp_small_packet(ws,result)
        elif "航空包裹" in sn: _cp_parcel(ws,"国际航空包裹",result)
        elif "SAL" in sn: _cp_parcel(ws,"国际SAL包裹",result)
        elif "水陆路" in sn: _cp_parcel(ws,"国际水陆路包裹",result)
        elif "e包裹" in sn: _cp_parcel(ws,"e包裹",result)
        elif "港澳台" in sn: _cp_parcel(ws,"港澳台包裹",result)
    return result

def _cp_ems(ws, result):
    for ri in range(5, min(200, ws.max_row+1)):
        ct = ss(cell(ws,ri,2)); dp = sf(cell(ws,ri,3)); ip = sf(cell(ws,ri,4)); cp = sf(cell(ws,ri,5))
        if not ct or (dp is None and ip is None): continue
        for c in split_countries(ct):
            for pt,pr in [("文件",dp),("物品",ip)]:
                if pr is None: continue
                name = f"EMS-{pt}-{c}"
                result["routes"].append(new_route("中国邮政",name,c,"","express","干线+尾程双清派送",
                    door_delivery=True, customs_mode="DDP"))
                result["pricing"].append(new_pricing(name,
                    first_weight_price=pr, first_weight_unit="500g",
                    continued_weight_price=cp, continued_weight_unit="500g"))

def _cp_e_express(ws, result):
    for ri in range(6, min(150, ws.max_row+1)):
        c = ss(cell(ws,ri,2)); fp = sf(cell(ws,ri,3)); cp = sf(cell(ws,ri,4)); mx = sf(cell(ws,ri,5))
        if not c or fp is None: continue
        name = f"e特快-{c}"
        result["routes"].append(new_route("中国邮政",name,c,"","express","干线+尾程双清派送",
            door_delivery=True, customs_mode="DDP"))
        result["pricing"].append(new_pricing(name,
            first_weight_price=fp, first_weight_unit="50g",
            continued_weight_price=cp, continued_weight_unit="50g", max_weight=mx))

def _cp_eub(ws, sn, result):
    pn = "e邮宝特惠" if "特惠" in sn else "e邮宝"
    for ri in range(5, min(100, ws.max_row+1)):
        c = ss(cell(ws,ri,2)); pi = sf(cell(ws,ri,4)); pk = sf(cell(ws,ri,5))
        mn = sf(cell(ws,ri,6)); mx = sf(cell(ws,ri,7))
        if not c or (pi is None and pk is None): continue
        name = f"{pn}-{c}"
        result["routes"].append(new_route("中国邮政",name,c,"","postal","干线+尾程双清派送",
            door_delivery=True, customs_mode="DDP"))
        result["pricing"].append(new_pricing(name,
            min_weight=mn/1000 if mn else 0, max_weight=mx/1000 if mx else None,
            registration_fee=pi, continued_weight_price=pk))

def _cp_small_packet(ws, result):
    for ri in range(6, min(250, ws.max_row+1)):
        c = ss(cell(ws,ri,2))
        opi = sf(cell(ws,ri,3)); opk = sf(cell(ws,ri,4))
        rpi = sf(cell(ws,ri,5)); rpk = sf(cell(ws,ri,6))
        if not c: continue
        if opi is not None or opk is not None:
            name = f"国际小包-平常-{c}"
            result["routes"].append(new_route("中国邮政",name,c,"","postal","干线+尾程双清派送",
                door_delivery=True, customs_mode="DDP", supported_categories="普货"))
            result["pricing"].append(new_pricing(name,
                min_weight=0, max_weight=2, registration_fee=opi, continued_weight_price=opk))
        if rpi is not None or rpk is not None:
            name = f"国际小包-挂号-{c}"
            result["routes"].append(new_route("中国邮政",name,c,"","postal","干线+尾程双清派送",
                door_delivery=True, customs_mode="DDP", supported_categories="普货"))
            result["pricing"].append(new_pricing(name,
                min_weight=0, max_weight=2, registration_fee=rpi, continued_weight_price=rpk))

def _cp_parcel(ws, pn, result):
    for ri in range(4, min(200, ws.max_row+1)):
        c = ss(cell(ws,ri,2)); fp = sf(cell(ws,ri,3)); cp = sf(cell(ws,ri,4))
        if not c or fp is None: continue
        name = f"{pn}-{c}"
        result["routes"].append(new_route("中国邮政",name,c,"","postal","干线+尾程双清派送",
            door_delivery=True, customs_mode="DDP"))
        result["pricing"].append(new_pricing(name,
            first_weight_price=fp, first_weight_unit="KG",
            continued_weight_price=cp, continued_weight_unit="KG"))


# ============================================================
# 主入口
# ============================================================

def main():
    p = argparse.ArgumentParser(description="跨境物流报价表解析器 v4 - 支持20种货代格式")
    p.add_argument("--input", required=True, help="Excel文件路径")
    p.add_argument("--output", default="", help="输出JSON文件路径")
    p.add_argument("--supplier", default="", help="强制指定货代类型")
    args = p.parse_args()

    if not os.path.exists(args.input):
        print(json.dumps({"error":f"文件不存在: {args.input}"},ensure_ascii=False)); sys.exit(1)

    # .xls files
    if args.input.endswith('.xls') and not args.input.endswith('.xlsx'):
        fn = os.path.basename(args.input)
        supplier = args.supplier or detect_supplier(fn, [])
        if supplier == "启丰物流":
            result = parse_qifeng_xls(args.input)
        elif supplier == "昌海运通":
            result = parse_changhai(args.input, None)
        else:
            print(json.dumps({"error":f".xls格式暂不支持该货代: {supplier}"},ensure_ascii=False))
            sys.exit(1)
    else:
        wb = openpyxl.load_workbook(args.input, data_only=True)
        supplier = args.supplier or detect_supplier(os.path.basename(args.input), wb.sheetnames)

        parsers = {
            "蓝牛国际": parse_lanniu,
            "亿俐缇": parse_yiliti,
            "福鑫国际": parse_fuxin,
            "韩润物流": parse_hanrun,
            "祥源国际": parse_xiangyuan,
            "畅骏国际": parse_changjun,
            "启丰物流": parse_qifeng,
            "海源日本": parse_haiyuan,
            "卉马国际": parse_huima,
            "昌海运通": lambda f,w: parse_changhai(f,w),
            "翔平": parse_xiangping,
            "艾姆勒": parse_iml,
            "惠运捷邮政大包": parse_hyj_postal_big,
            "满天星综合": parse_manTianXing_comprehensive,
            "满天星主营": parse_manTianXing_main,
            "惠运捷": parse_huiYunJie,
            "元昊": parse_yuanHao,
            "中国邮政": parse_chinaPost,
        }
        if supplier not in parsers:
            print(json.dumps({"error":f"无法识别货代格式: {supplier}",
                             "sheets": wb.sheetnames},ensure_ascii=False))
            sys.exit(1)
        result = parsers[supplier](args.input, wb)
        wb.close()

    stats = {k: len(v) for k,v in result.items() if k!="parse_time" and isinstance(v,list)}
    stats["parse_time"] = result["parse_time"]
    result["stats"] = stats

    out = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        with open(args.output,'w',encoding='utf-8') as f: f.write(out)
        print(json.dumps({"status":"success","output":args.output,"stats":stats},ensure_ascii=False))
    else:
        print(out)

if __name__ == "__main__":
    main()
